# -*- coding: utf-8 -*-
"""
@Author  : XIAOAN
@Time    : 2025/2/11 21:45
@File    : excel_reader.py
@Description : 
@progname: xz_api_auto	
"""

import openpyxl

def read_appLoginMsg(file_path, sheet_name):
    init_msg_app = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    init_msg_app.append(sheet["C2"].value)
    init_msg_app.append(sheet["D2"].value)
    init_msg_app.append(sheet["E2"].value)
    return init_msg_app
    # 首先将登录用例单独出来，用来获取bearer值
    # if sheet_name == "application":
    #     return init_msg_app
    # elif sheet_name == "system":
    #     login_url_sys = sheet["C2"].value
    #     login_method_sys = sheet["D2"].value
    #     login_param_sys = sheet["E2"].value
    #     init_msg_sys.append(login_url_sys)
    #     init_msg_sys.append(login_method_sys)
    #     init_msg_sys.append(login_param_sys)
    #     return init_msg_sys

def read_sysLoginMsg(file_path, sheet_name):
    init_msg_sys = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    # 首先将登录用例单独出来，用来获取bearer值
    init_msg_sys.append(sheet["C2"].value)
    init_msg_sys.append(sheet["D2"].value)
    init_msg_sys.append(sheet["E2"].value)
    return init_msg_sys

def read_app_test_cases(file_path, sheet_name):
    """读取 Excel 文件中的测试用例"""
    test_cases_app = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(min_row=3, values_only=True):  # 从第三行开始读取
        # 读取用例数据
        test_case_app = {
            "case_name": row[1],  # 用例名称
            "url": row[2],  # 请求URL
            "method": row[3],     # 请求方法
            "params": eval(row[4]) if row[4] else "",   # 请求参数
            "expected_status": row[5],  # 预期状态码
            "assert_key": row[6],       # 断言字段
            "assert_value": row[7]      # 断言值
        }
        test_cases_app.append(test_case_app)

    return test_cases_app

def read_sys_test_cases(file_path, sheet_name):
    """读取 Excel 文件中的测试用例"""
    test_cases_sys = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(min_row=3, values_only=True):  # 从第三行开始读取
        # 读取用例数据
        test_case_sys = {
            "case_name": row[1],  # 用例名称
            "url": row[2],  # 请求URL
            "method": row[3],     # 请求方法
            "params": eval(row[4]) if row[4] else "",   # 请求参数
            "expected_status": row[5],  # 预期状态码
            "assert_key": row[6],       # 断言字段
            "assert_value": row[7]      # 断言值
        }
        test_cases_sys.append(test_case_sys)

    return test_cases_sys