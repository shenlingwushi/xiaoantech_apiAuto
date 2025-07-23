# -*- coding: utf-8 -*-
"""
@Author  : XIAOAN
@Time    : 2025/2/11 17:03
@File    : getExcel.py
@Description : 
@progname: xz_api_auto	
"""

import os
from openpyxl import load_workbook
from common.public_path import DIR
# from common.config_operate_api import Config

class GetExcelData():
    """
    封装读取Excel数据
    """

    def __init__(self,sheet):
        self.path = r"C:\Users\XIAOAN\PycharmProjects\xz_api_auto\test_data\test.xlsx" #excel文件路径
        self.excel_path = os.path.join(DIR,self.path)
        self.wb = load_workbook(self.excel_path)
        self.ws = self.wb[sheet]
        self.max_columns = self.ws.max_column  #最大列数
        self.max_rows=self.ws.max_row  #最大行数

    def get_row_case_list(self,row=None):
        """
        按行获取Excel中的用例数据，返回list,如果row=None时，返回整个sheet页所有数据（除表头），
        如果row为具体数字时，读取sheet对应的行数数据
        :param row: 行数，第一行数据为title，默认已把值加1
        :return:
        """
        case_list = [] #返回的所有case数据
        #当row为None返回当前sheet页中所有用例数据
        if row==None:
            for i in range(self.max_rows):
                temp_case_list=[]
                for each in self.ws.iter_cols(min_col=0):
                    temp_case_list.append(each[i].value)
                #openpy的iter_cols用法会读取所有行包含空行（做了格式其他的改变，也会读取），加判断去除空行
                if temp_case_list[0]!=None and temp_case_list[:-1]!=None:
                    case_list.append(temp_case_list)
            #去除表头数据
            del case_list[0]
            return case_list
        else:
            for i in range(1,self.max_columns+1):
                value=self.ws.cell(row=row+1,column=i).value
                case_list.append(value)
            return case_list

    def get_row_case_dict(self,row=None):
        """
        按行获取Excel中的用例数据，如果row=None时，返回的数据是全部用例数据，格式为list中存在多个dict
        如果row等于具体数字时，读取对应行的数据
        :param row: 行数
        :return:
        """
        case_title_list=self.get_row_case_list(row=0) #获取sheet页第一行，即title
        if row==None:
            all_case_dict_list=[]  #存每个用例的dict格式的list
            all_case_list = self.get_row_case_list()
            for case in all_case_list:
                temp_case_dict=dict(zip(case_title_list,case))
                all_case_dict_list.append(temp_case_dict)
            return all_case_dict_list
        else:
            case_list=self.get_row_case_list(row=row)
            #通过title和一行的数据使用zip合并成字典
            case_dict=dict(zip(case_title_list,case_list))
            return case_dict


    def get_case_data(self,row=None):
        """
        按行获取Excel中用例数据，并把数据中提取url、data、expected_result值，
        返回tuple,其中从Excel中读取的键值对数据需要用eval格式转成字典格式
        row==None时返回全部用例数据
        :param row: 行数
        :return:
        """

        if row==None:
            all_case_list = [] #list存多个tuple，每个tuple中有url，data，expected_result
            all_case_dict_list = self.get_row_case_dict()
            for temp_case_dict in all_case_dict_list:
                temp_list=[]
                temp_list.append(temp_case_dict["url"])
                data = temp_case_dict["data"]
                temp_list.append(eval(data))
                temp_list.append(temp_case_dict["expected_result"])
                all_case_list.append(tuple(temp_list))
            return all_case_list
        else:
            case_dict=self.get_row_case_dict(row=row)
            new_case_list=[]
            new_case_list.append(case_dict["url"])
            data=case_dict["data"]
            new_case_list.append(eval(data))
            new_case_list.append(case_dict["expected_result"])

            return tuple(new_case_list)
